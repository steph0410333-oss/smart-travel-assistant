from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROFILE_SHEET = "星期時段擁擠度"
DETAIL_SHEET = "歷史人流明細"
METHOD_SHEET = "方法說明"
MODEL_VERSION = "crowd_40_40_10_10_v1"
MINIMUM_SAMPLE_COUNT = 8
EXPECTED_LEVELS = {"舒適", "普通", "偏擠", "擁擠", "非常擁擠"}
REQUIRED_PROFILE_COLUMNS = {
    "station",
    "weekday_num",
    "hour",
    "sample_count",
    "final_crowd_score",
    "final_crowd_level",
    "score_status",
}


def _iso_date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return date.fromisoformat(str(value)[:10]).isoformat()


def _source_checksum(source: Path) -> str:
    digest = hashlib.sha256()
    with source.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _detail_metadata(workbook) -> dict[str, Any]:
    sheet = workbook[DETAIL_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    index = {name: headers.index(name) for name in headers}
    required = {"date", "station"}
    missing = required - set(index)
    if missing:
        raise ValueError(f"{DETAIL_SHEET} 缺少欄位：{', '.join(sorted(missing))}")

    earliest: str | None = None
    latest: str | None = None
    months: set[str] = set()
    stations: set[str] = set()
    row_count = 0
    for row in rows:
        if row[index["date"]] is None or row[index["station"]] is None:
            continue
        current_date = _iso_date(row[index["date"]])
        earliest = current_date if earliest is None or current_date < earliest else earliest
        latest = current_date if latest is None or current_date > latest else latest
        months.add(current_date[:7])
        stations.add(str(row[index["station"]]))
        row_count += 1

    return {
        "data_period_start": earliest,
        "data_period_end": latest,
        "source_month_count": len(months),
        "station_count": len(stations),
        "station_hour_rows": row_count,
    }


def _method_metadata(workbook) -> dict[str, str]:
    sheet = workbook[METHOD_SHEET]
    return {
        str(item): str(description)
        for item, description in sheet.iter_rows(min_row=2, values_only=True)
        if item is not None and description is not None
    }


def build_payload(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    expected_sheets = {PROFILE_SHEET, DETAIL_SHEET, METHOD_SHEET}
    missing_sheets = expected_sheets - set(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"來源活頁簿缺少工作表：{', '.join(sorted(missing_sheets))}")

    metadata = _detail_metadata(workbook)
    method = _method_metadata(workbook)
    sheet = workbook[PROFILE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    index = {name: headers.index(name) for name in headers}
    missing_columns = REQUIRED_PROFILE_COLUMNS - set(index)
    if missing_columns:
        raise ValueError(f"{PROFILE_SHEET} 缺少欄位：{', '.join(sorted(missing_columns))}")

    stations: dict[str, dict[str, dict[str, list[Any]]]] = {}
    profile_count = 0
    unavailable_count = 0
    for row_number, row in enumerate(rows, start=2):
        station_value = row[index["station"]]
        score_value = row[index["final_crowd_score"]]
        status = str(row[index["score_status"]] or "資料不足")
        if station_value is None:
            continue
        if status != "正常" or score_value is None:
            unavailable_count += 1
            continue

        score = float(score_value)
        level = str(row[index["final_crowd_level"]])
        if not 0 <= score <= 100:
            raise ValueError(f"{PROFILE_SHEET}!第{row_number}列分數超出0–100：{score}")
        if level not in EXPECTED_LEVELS:
            raise ValueError(f"{PROFILE_SHEET}!第{row_number}列等級無效：{level}")

        station = str(station_value)
        weekday = str(int(row[index["weekday_num"]]))
        hour = str(int(row[index["hour"]]))
        sample_count = int(row[index["sample_count"]])
        profile = [round(score, 2), sample_count, level, status]
        station_hours = stations.setdefault(station, {}).setdefault(weekday, {})
        if hour in station_hours:
            raise ValueError(f"重複查詢鍵：{station}/{weekday}/{hour}")
        station_hours[hour] = profile
        profile_count += 1

    if profile_count == 0:
        raise ValueError("來源活頁簿沒有可用的人流分數")
    if len(stations) != metadata["station_count"]:
        raise ValueError(
            f"工作表站數不一致：{PROFILE_SHEET}={len(stations)}、"
            f"{DETAIL_SHEET}={metadata['station_count']}"
        )

    return {
        "metadata": {
            **metadata,
            "source_workbook": source.name,
            "source_sha256": _source_checksum(source),
            "profile_sheet": PROFILE_SHEET,
            "detail_sheet": DETAIL_SHEET,
            "method_sheet": METHOD_SHEET,
            "model_name": "歷史相對人流擁擠指數（40/40/10/10）",
            "model_version": MODEL_VERSION,
            "lookup_key": ["station", "weekday_num", "hour"],
            "profile_encoding": [
                "final_crowd_score",
                "sample_count",
                "final_crowd_level",
                "score_status",
            ],
            "weights": {
                "within_day_total_score": 0.40,
                "across_weekday_total_score": 0.40,
                "entry_pressure_score": 0.10,
                "exit_pressure_score": 0.10,
            },
            "level_thresholds": {
                "舒適": [0, 35],
                "普通": [35, 55],
                "偏擠": [55, 70],
                "擁擠": [70, 85],
                "非常擁擠": [85, 100],
            },
            "calculation_formula": method.get("最終公式"),
            "minimum_sample_count": MINIMUM_SAMPLE_COUNT,
            "profile_count": profile_count,
            "unavailable_profile_count": unavailable_count,
            "data_label": "歷史OD相對人流推估，非即時站內人數",
            "reliability_label": "experimental",
            "limitations": [
                "目前僅有一個月資料，同星期同時段通常只有4至5筆樣本。",
                "final_crowd_score是歷史相對排名分數，不是官方容量使用率。",
                "進站加出站代表交通活動量，不是站內同時存在人數。",
            ],
        },
        "stations": stations,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="將40/40/10/10人流Excel轉為APP用JSON")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    payload = build_payload(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"輸出模型 {payload['metadata']['model_version']}："
        f"{payload['metadata']['profile_count']} 筆、"
        f"{len(payload['stations'])} 站至 {args.output}"
    )


if __name__ == "__main__":
    main()
